from DrissionPage import ChromiumPage,ChromiumOptions
from openpyxl import load_workbook, Workbook
from urllib.parse import quote

import time
import hashlib
data = b'Hello, World!'
hash_object = hashlib.sha1(data)
hexdigest = hash_object.hexdigest()
import os
import signal
import datetime
from metaData import boss_data
# 薪资 - salary
"""
不限：''
3k以下：'402'
3-5k：'403'
5-10k：'404'
10-20k：'405'
20-50k：'406'
50k以上：'407'
"""

# 工作年限 - job_experience
"""
不限：''
在校生：'108'
应届生：'102'
经验不限：'101'
1年以内：'103'
1-3年：'104'
3-5年：'105'
5-10年：'106'
10年以上：'107'
"""
# 复合案例：'102,103'

# 学历 - job_degree
"""
不限：''
初中及以下：'209'
中技/中专：'208'
高中：'206'
大专：'202'
本科：'203'
硕士：'204'
博士：'205'
"""
# 复合案例：'203,204'

class Boss():
    def generate_query_urls(self,city,job_name,job_salarys=['405']):
        city_num = boss_data.citys[city]
        url_list = []
        for job_salary in job_salarys:
            url_list.append(f"https://www.zhipin.com/web/geek/job?query={quote(job_name)}&city={city_num}&salary={job_salary}")

        for first_level_place in boss_data.city_info[city]:
            first_level_code = first_level_place["code"]
            for job_salary in job_salarys:
                url_list.append(
                    f"https://www.zhipin.com/web/geek/job?query={quote(job_name)}&city={city_num}&salary={job_salary}&areaBusiness={first_level_code}")
                if "subLevelModelList" in first_level_place:
                    sub_level_Model_list = first_level_place["subLevelModelList"]
                    for sub_level_model in sub_level_Model_list:
                            second_level_code=sub_level_model["code"]
                            url_list.append(
                                    f"https://www.zhipin.com/web/geek/job?query={quote(job_name)}&city={city_num}&salary={job_salary}&areaBusiness={first_level_code}:{second_level_code}")
        return url_list
    def __init__(self,job_name='',job_salary=['405'],city='chengdu',job_experience='0',job_degree='0',
        ignore_job=[],ignore_edu=[],ignore_area=[],ignore_salary=[],ignore_job_require=[],
        overlap=False,login_in=False):
        self.job_records = []
        self.login_in = login_in
        self.ignore_job = ignore_job
        self.ignore_edu = ignore_edu
        self.ignore_area = ignore_area
        self.ignore_salary = ignore_salary
        self.ignore_job_require = ignore_job_require
        self.city = city
        self.query_list = self.generate_query_urls(city,job_name,job_salary)

        file_name = 'output/boss_crawl_output.xlsx'

        if not os.path.exists('output'): os.mkdir('output')
        if overlap:
            counter = 1
            while os.path.exists(file_name):
                file_name = f'output/boss_crawl_output_{counter}.xlsx'
                counter += 1
        elif os.path.exists(file_name): os.remove(file_name)

        self.file_path = file_name

        wb = Workbook()
        ws = wb.active
        row = ws.max_row

        ws.cell(row=row, column=1, value='职位')
        ws.cell(row=row, column=2, value='工资')
        ws.cell(row=row, column=3, value='公司名')
        ws.cell(row=row, column=4, value='位置')
        ws.cell(row=row, column=5, value='详细内容')
        ws.cell(row=row, column=6, value='招聘页面')
        ws.cell(row=row, column=7, value='Boss活跃时间')

        wb.save(self.file_path)
        wb.close()

    def total_run(self):
        now_num = 0
        while self.query_list and len(self.query_list) > 0:
            url = self.query_list[0]
            self.query_list = self.query_list[1:]
            now_num += 1
            print("remain urls are %s, now  crawled url %s........." % (len(self.query_list), now_num))
            current_time = datetime.datetime.now()
            print("当前时间：", current_time)
            print(url)
            self.run(url)

    def run(self,page_url):
        # search_timeout必需大于或等于0.3
        search_timeout = 0.3
        try:
            page = ChromiumPage(ChromiumOptions().headless())
        except Exception as e:
            print(e)

        # APP扫码登录
        if self.login_in:
            print("login.....")
            page.get('https://www.zhipin.com/')

            if not page.ele('x://a[@ka="header-username"]',timeout=search_timeout):
                page.ele('x://a[contains(@class,"header-login-btn")]',timeout=search_timeout).click()
                page.ele('x://div[contains(@class,"btn-sign-switch")]',timeout=search_timeout).click()

            while not page.ele('x://a[@ka="header-username"]',timeout=search_timeout):
                try: page = page.latest_tab
                except: pass
                time.sleep(1)
            print("login end")

        page.get(page_url)
        try:
            while True:
                time.sleep(0.15)
                # 屏蔽登录窗口
                try: page.run_js('Object.defineProperties(navigator,{webdriver:{get:()=>false}})')
                except: pass
                pages_string = page.ele('x://div[@class="options-pages"]').text
                if "10" not in pages_string:
                    #这种情况当前已经是足够了，不用细分了.把更细的url都删掉
                    new_list = []
                    for url in self.query_list:
                        if page_url not in url:
                            new_list.append(url)
                    new_list.sort(key=len)
                    self.query_list = new_list
                else:
                    for url in self.query_list:
                        if page_url in url:
                            #说明还有比这个链接更细的链接，这个链接就不用爬了
                            raise Exception("还有更细链接,这个链接不用爬")


                box = page.ele('x://ul[@class="job-list-box"]').children(timeout=search_timeout)
                for card in box:
                    if not page.ele('x://a[@ka="header-username"]', timeout=search_timeout):
                        # 关闭登录窗口
                        try: page.ele('x://span[@class="boss-login-close"]',timeout=4).click()
                        except: pass

                    continue_flag = False
                    # 职位
                    job_name = card.ele('x://span[@class="job-name"]',timeout=search_timeout).text
                    for each in self.ignore_job:
                        if each in job_name:
                            continue_flag = True
                            break
                    if continue_flag: continue

                    # 薪资
                    job_salary = card.ele('x://span[@class="salary"]',timeout=search_timeout).text
                    job_salary = f'[{job_salary}]'
                    for each in self.ignore_salary:
                        if each in job_salary:
                            continue_flag = True
                            break
                    if continue_flag: continue

                    # 学历
                    job_tags = ''
                    try:
                        for each in card.ele('x://ul[@class="tag-list"]',timeout=search_timeout).children(timeout=search_timeout):
                            job_tags = f'{job_tags}/{each.text}'
                    except: pass

                    for each in self.ignore_edu:
                        if each in job_tags:
                            continue_flag = True
                            break
                    if continue_flag: continue

                    # 地区
                    job_area = card.ele('x://span[@class="job-area"]',timeout=search_timeout).text
                    for each in self.ignore_area:
                        if each in job_area:
                            continue_flag = True
                            break
                    if continue_flag: continue

                    job_company = card.ele('x://h3[@class="company-name"]',timeout=search_timeout).child(timeout=search_timeout).text
                    job_company_tags = []
                    try:
                        for each in card.ele('x://ul[@class="company-tag-list"]',timeout=search_timeout).children(timeout=search_timeout):
                            job_company_tags.append(each.text)
                    except: pass

                    try:
                        job_content = ''
                        job_link = ''
                        card.ele('x://span[@class="job-name"]',timeout=search_timeout).click()
                        tab = page.latest_tab

                        try:
                            job_link = tab.url

                            time.sleep(0.5)
                            tab.scroll.down(700)
                            time.sleep(1)

                            try:
                                try:
                                    job_info_tag = tab.ele('x://div[contains(@class,"job-tags")]',timeout=search_timeout)

                                    for each in job_info_tag.children(timeout=search_timeout):
                                        job_content = f'{job_content}/{each.raw_text}'
                                except: pass

                                if job_content: job_content += '\n'
                                job_content += tab.ele('x://div[@class="job-sec-text"]',timeout=search_timeout).raw_text

                                for each in self.ignore_job_require:
                                    if each in job_content:
                                        continue_flag = True
                                        break
                            except: pass

                            if continue_flag:
                                tab.close()
                                continue
                            try:
                                try:
                                    boss_active_time = tab.ele('x://span[@class="boss-active-time"]',timeout=search_timeout).text
                                except:pass
                                if not boss_active_time:
                                    boss_active_time = "未知"
                            except:pass

                            message = f'职位:{job_name},工资:{job_salary},公司名:{job_company},位置:{job_area},详细内容:\n{job_content},Boss活跃时间:{boss_active_time},招聘页面:{job_link}\n'
                            print(message)

                            #防止重复存储
                            hash_object = hashlib.sha1(message)
                            hexdigest = hash_object.hexdigest()
                            if hexdigest in self.job_records:
                                continue
                            else:
                                self.job_records.append(hexdigest)

                            try:
                                wb = load_workbook(self.file_path)
                                ws = wb.active
                                row = ws.max_row + 1

                                ws.cell(row=row, column=1, value=job_name)
                                ws.cell(row=row, column=2, value=job_salary)
                                ws.cell(row=row, column=3, value=job_company)
                                ws.cell(row=row, column=4, value=job_area)
                                ws.cell(row=row, column=5, value=job_content)
                                ws.cell(row=row, column=6, value=job_link)
                                ws.cell(row=row, column=7, value=boss_active_time)

                                wb.save(self.file_path)
                                wb.close()
                            except Exception as e:
                                print("save job Error:")
                                print(e)
                        except: pass
                        tab.close()
                    except: pass

                next_page = page.ele('x://i[@class="ui-icon-arrow-right"]')
                if 'disabled' == next_page.parent().attr('class'):
                    print("Boss直聘 Crawl End")
                    break
            next_page.click()
            time.sleep(0.5)
        except Exception as e:
            print(e)
            print("while circle error")
        page.close()


if __name__ == '__main__':
    try:
        boss = Boss(
            job_name='',
            job_salary='405',
            job_experience='102',
            job_degree='203',
            ignore_job=['教师', '老师', '销售', '保险', '消防', '前台', '客服', '营销', '售后', '管培生', '助理','经理', '硕士', '博士', '商务'],
            ignore_edu=['急聘', '1-3年', '3-5年', '5-10年', '2个月', '1年以下', '3年以上', '10年以上'],
            ignore_area=['北京', '上海'],
            ignore_salary=['面议', '[0-', '[1-', '[2-', '[3-', '[4-', '[5-', '[6-', '[7-', '[8-', '[9-', '100-250元/天', '200-300元/天'],
            ignore_job_require=['硕士学历', '硕士及以上学历', '硕士以上学历', '博士学历', '博士及以上学历', '博士以上学历', '985', '211', '出差'],
            login_in=True,
            overlap=False
        )
        boss.run()
    except: os.kill(os.getpid(), signal.SIGINT)