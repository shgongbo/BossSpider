from DrissionPage import ChromiumPage
from openpyxl import load_workbook, Workbook
from urllib.parse import quote

import time
import os
import signal

class NiuKe():
    def __init__(self,job_name = 'C++后端',overlap=False):
        self.query = f'https://www.nowcoder.com/jobs/school/jobs?search={quote(job_name)}&order=1'
        file_name = '../output/niuke_crawl_output.xlsx'

        if not os.path.exists('../output'): os.mkdir('../output')
        if overlap:
            counter = 1
            while os.path.exists(file_name):
                file_name = f'../output/niuke_crawl_output_{counter}.xlsx'
                counter += 1
        elif os.path.exists(file_name):
            os.remove(file_name)

        self.file_path = file_name

        wb = Workbook()
        ws = wb.active
        row = ws.max_row

        ws.cell(row=row, column=1, value='职位')
        ws.cell(row=row, column=2, value='工资')
        ws.cell(row=row, column=3, value='公司名')
        ws.cell(row=row, column=4, value='位置')
        ws.cell(row=row, column=5, value='详细内容')
        ws.cell(row=row, column=6, value='招聘页面')

        wb.save(self.file_path)
        wb.close()
    def run(self):
        page = ChromiumPage()
        page.get(self.query)

        while True:
            time.sleep(0.1)
            # 多此一举，屏蔽登录窗口
            try: page.run_js('Object.defineProperties(navigator,{webdriver:{get:()=>false}})')
            except: pass

            box = page.eles('x://div[@class="job-card-item"]')
            for card in box:
                job_name_label = card.ele('x://span[@class="job-name"]')
                job_name = job_name_label.text
                job_salary = card.ele('x://span[contains(@class,"job-salary")]').text
                job_company = card.ele('x://div[@class="company-name"]').text
                job_area = ''

                job_content = ''
                job_link = ''
                try:
                    job_name_label.click()
                    tab = page.latest_tab

                    job_link = tab.url
                    for each in tab.ele('x://div[contains(@class,"job-detail-infos")]').children():
                        if each.children(timeout=0.2):
                            for key in each.children():
                                job_content = f'{job_content}{key.raw_text}'
                        else: job_content = f'{job_content}{each.raw_text}'

                    print(f'职位:{job_name},工资:{job_salary},公司名:{job_company},位置:,详细内容:{job_content},招聘页面:{job_link}\n')
                    time.sleep(0.35)
                    tab.close()
                except: pass
                time.sleep(0.1)
                wb = load_workbook(self.file_path)
                ws = wb.active
                row = ws.max_row + 1

                ws.cell(row=row, column=1, value=job_name)
                ws.cell(row=row, column=2, value=job_salary)
                ws.cell(row=row, column=3, value=job_company)
                ws.cell(row=row, column=4, value=job_area)
                ws.cell(row=row, column=5, value=job_content)
                ws.cell(row=row, column=6, value=job_link)

                wb.save(self.file_path)
                wb.close()

            next_page = page.ele('x://button[@class="btn-next"]')
            if 'disabled' == next_page.attr('disabled'):
                print('牛客网 Crwal End')
                break
            next_page.click()

        page.close()


if __name__ == '__main__':
    try:
        niuke = NiuKe(job_name='C++')
        niuke.run()
    except: os.kill(os.getpid(), signal.SIGINT)
