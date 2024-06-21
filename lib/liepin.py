from DrissionPage import ChromiumPage
from urllib.parse import quote
from openpyxl import load_workbook, Workbook

import time
import os
import signal

# 薪资 - salary
"""
 自定义
 nk-Nk: 'n$N', 例如 11k-15k - '11$15'
"""

# 工作年限 - job_experience
"""
不限：''
应届生：'1'
实习生：'2'
1年以内：'0$1'
1-3年：'1$3'
3-5年：'3$5'
5-10年：'5$10'
10年以上：'10$999'
"""

# 学历 - job_degree
"""
博士：'010'
MBA/EMBA：'020'
硕士：'030'
本科：'040'
大专：'050'
中专/中技：'060'
高中：'080'
初中及以下：'090'
"""

class LiePin():
    def __init__(self,job_name='C++后端',pubTime='7',job_salary='10&20',job_experience='0',job_degree='0',
        ignore_job=[],ignore_edu=[],ignore_area=[],ignore_salary=[],ignore_job_require=[],
        overlap=False,login_in=False):
        self.login_in = login_in
        self.ignore_job = ignore_job
        self.ignore_edu = ignore_edu
        self.ignore_area = ignore_area
        self.ignore_salary = ignore_salary
        self.ignore_job_require = ignore_job_require

        # pubTime - '1', '3', '7', '30',n天以内
        pageSize = 40
        self.query = f'https://www.liepin.com/zhaopin/?&currentPage=0&city=410&dq=410&pubTime={pubTime}&pageSize={pageSize}&key={quote(job_name)}&suggestTag=&workYearCode={job_experience}&compId=&compName=&compTag=&compStage=&compKind=&compScale=&industry=&salary={job_salary}&jobKind=&eduLevel={job_degree}'

        file_name = '../output/liepin_crawl_output.xlsx'

        if not os.path.exists('../output'): os.mkdir('../output')
        if overlap:
            counter = 1
            while os.path.exists(file_name):
                file_name = f'../output/liepin_crawl_output_{counter}.xlsx'
                counter += 1
        elif os.path.exists(file_name): os.remove(file_name)

        self.file_path = file_name

        wb = Workbook()

        ws = wb.active
        row = ws.max_row

        ws.cell(row=row, column=1, value='职位')
        ws.cell(row=row, column=2, value='工资')
        ws.cell(row=row, column=3, value='公司名')
        ws.cell(row=row, column=4, value='位置')
        ws.cell(row=row, column=5, value='详细内容')
        ws.cell(row=row, column=6, value='招聘页面')

        wb.save(self.file_path)
        wb.close()

    def run(self):
        # search_timeout必需大于0.2
        search_timeout = 0.25
        
        page = ChromiumPage()

        # APP扫码登录
        if self.login_in:
            print("login.....")
            page.get('https://www.liepin.com/')

            if not page.ele('x://div[@id="header-quick-menu-user-info"]',timeout=search_timeout):
                page.ele('x://div[contains(@class,"btn-sign-switch")]').click()
                page.ele('x://div[contains(@class,"login-tabs-line-text")]').click()

            while not page.ele('x://div[@id="header-quick-menu-user-info"]',timeout=search_timeout):
                try: page = page.latest_tab
                except: pass
                time.sleep(1)
            print("login end")

        page.get(self.query)

        while True:
            time.sleep(0.25)
            # 多此一举，屏蔽滑动验证
            try: page.run_js('Object.defineProperties(navigator,{webdriver:{get:()=>false}})')
            except: pass

            box = page.ele('x://div[@class="job-list-box"]').children(timeout=search_timeout)
            for card in box:
                continue_flag = False
                # 职位
                job_name = card.ele('x://div[contains(@class,"job-title")]',timeout=search_timeout).child(timeout=search_timeout).text
                for each in self.ignore_job:
                    if each in job_name:
                        continue_flag = True
                        break
                if continue_flag: continue

                # 薪资
                job_salary = card.ele('x://span[contains(@class,"job-salary")]',timeout=search_timeout).text
                job_salary = f'[{job_salary}]'
                for each in self.ignore_salary:
                    if each in job_salary:
                        continue_flag = True
                        break
                if continue_flag: continue

                # 学历
                job_tags = ''
                try:
                    addtion_tag = card.ele('x://span[contains(@class,"job-tag")]', timeout=search_timeout).text
                    job_tags = f'{job_tags}/{addtion_tag}'
                except: pass

                try:
                    for each in card.ele('x://div[contains(@class,"job-labels-box")]',timeout=search_timeout).children(timeout=search_timeout):
                        job_tags = f'{job_tags}/{each.text}'
                except: pass

                for each in self.ignore_edu:
                    if each in job_tags:
                        continue_flag = True
                        break
                if continue_flag: continue

                # 地区
                job_area = ''
                try:
                    for each in card.ele('x://div[contains(@class,"job-dq-box")]',timeout=search_timeout).children(timeout=search_timeout):
                        job_area += each.text
                except: pass

                for each in self.ignore_area:
                    if each in job_area:
                        continue_flag = True
                        break
                if continue_flag: continue

                job_company = card.ele('x://span[contains(@class,"company-name")]',timeout=search_timeout).text
                job_company_tags = []
                try:
                    for each in card.ele('x://div[contains(@class,"company-tags-box")]',timeout=search_timeout).children(timeout=search_timeout):
                        job_company_tags.append(each.text)
                except: pass

                try:
                    job_content = ''
                    job_link = ''
                    card.ele('x://div[contains(@class,"job-detail-header-box")]',timeout=search_timeout).click()
                    tab = page.latest_tab

                    try:
                        time.sleep(0.5)
                        tab.scroll.down(400)
                        time.sleep(1)

                        job_info_content = tab.ele('x://section[@class="job-intro-container"]',timeout=search_timeout)
                        job_info_tag = tab.ele('x://div[@class="job-apply-container-left"]',timeout=search_timeout).child(timeout=search_timeout)
                        job_link = tab.url

                        for each in job_info_tag.children(timeout=search_timeout):
                            job_content = f'{job_content}/{each.raw_text}'

                        if job_content: job_content = f'{job_content}\n'
                        for each in job_info_content.children(timeout=search_timeout):
                            job_content = f'{job_content}{each.raw_text}\n'

                        for each in self.ignore_job_require:
                            if each in job_content:
                                continue_flag = True
                                break

                        time.sleep(1)
                        if continue_flag:
                            tab.close()
                            continue

                        print(f'职位:{job_name},工资:{job_salary},公司名:{job_company},位置:{job_area},详细内容:\n{job_content},招聘页面:{job_link}\n')
                        tab.close()

                        wb = load_workbook(self.file_path)
                        ws = wb.active
                        row = ws.max_row + 1

                        ws.cell(row=row, column=1, value=job_name)
                        ws.cell(row=row, column=2, value=job_salary)
                        ws.cell(row=row, column=3, value=job_company)
                        ws.cell(row=row, column=4, value=job_area)
                        ws.cell(row=row, column=5, value=job_content)
                        ws.cell(row=row, column=6, value=job_link)

                        wb.save(self.file_path)
                        wb.close()
                    except: pass
                    tab.close()
                except: pass

            next_page = page.ele('x://li[@title="Next Page"]')
            if 'false' != next_page.attr('aria-disabled'):
                print("猎聘 Crawl End")
                break
            next_page.click()
            time.sleep(0.5)
        page.close()


if __name__ == '__main__':
    try:
        liepin = LiePin(
            job_name='',
            pubTime='30',
            job_salary='10$20',
            job_experience='1',
            job_degree='040',
            ignore_job=['教师','老师','销售','保险','消防','前台','客服','营销','售后','管培生','助理','经理','硕士','博士','商务'],
            ignore_edu=['急聘','1-3年','3-5年','5-10年','2个月','1年以下','3年以上','10年以上'],
            ignore_area=['北京','上海'],
            ignore_salary=['面议','[0-','[1-','[2-','[3-','[4-','[5-','[6-','[7-','[8-','[9-','100-250元/天','200-300元/天'],
            ignore_job_require=['硕士学历','硕士及以上学历','硕士以上学历','博士学历','博士及以上学历','博士以上学历','985','211','出差'],
            login_in=True,
            overlap=False
        )
        liepin.run()
    except: os.kill(os.getpid(), signal.SIGINT)