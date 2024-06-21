"""
吐槽两句，
51job有一些莫名其妙的公司详细网址，
某些公司的详情网址还能自定义的，没有统一格式
"""

from DrissionPage import ChromiumPage
from DrissionPage.common import Actions
from openpyxl import load_workbook, Workbook
from urllib.parse import quote

import time
import random
import os
import signal

# 薪资 - job_salary
"""
所有：'0'
2k以下：'01'
2-3k：'02'
3-4.5k：'03'
4.5-6k：'04'
6-8k：'05'
0.8-1w：'06'
1-1.5w：'07'
1.5-2w：'08'
2-3w：'09'
3-4w：'10'
4-5w：'11'
5w以上：'12'
"""

# 工作年限 - job_experience
"""
所有：'0'
在校/应届：'01'
1-3年：'02'
3-5年：'03'
5-10年：'04'
10年以上：'05'
无需经验：'06'
"""

# 学历 - job_degree
"""
所有：'0'
初中及以下：'01'
高中/中技/中专：'02'
大专：'03'
本科：'04'
硕士：'05'
博士：'06'
无学历要求：'07'
"""

"""
使用登录前需将51job简历中心的简历删除
"""
class _51Job():
    def __init__(self,job_name='C++',job_salary='0',job_experience='0',job_degree='0',
        ignore_job=[],ignore_edu=[],ignore_area=[],ignore_salary=[],ignore_job_require=[],
        overlap=False,login_in=False):
        self.login_in = login_in
        self.ignore_job = ignore_job
        self.ignore_edu = ignore_edu
        self.ignore_area = ignore_area
        self.ignore_salary = ignore_salary
        self.ignore_job_require = ignore_job_require

        self.query = f'https://we.51job.com/pc/search?keyword={quote(job_name)}&searchType=2&sortType=1&salary={job_salary}&workYear={job_experience}&degree={job_degree}&metro='
        file_name = '../output/_51job_crawl_output.xlsx'

        if not os.path.exists('../output'): os.mkdir('../output')
        if overlap:
            counter = 1
            while os.path.exists(file_name):
                file_name = f'../output/_51job_crawl_output_{counter}.xlsx'
                counter += 1
        elif os.path.exists(file_name): os.remove(file_name)

        self.file_path = file_name

        wb = Workbook()
        ws = wb.active
        row = ws.max_row

        ws.cell(row=row, column=1, value='职位')
        ws.cell(row=row, column=2, value='工资')
        ws.cell(row=row, column=3, value='公司名')
        ws.cell(row=row, column=4, value='位置')
        ws.cell(row=row, column=5, value='详细内容')
        ws.cell(row=row, column=6, value='招聘页面')

        wb.save(self.file_path)
        wb.close()

    def run(self):
        # search_timeout必需大于0.2
        search_timeout = 0.25

        page = ChromiumPage()

        # 微信扫码登录
        if self.login_in:
            print("login.....")
            page.get('https://we.51job.com/')
            if not page.ele('x://a[@class="uname e_icon at"]',timeout=search_timeout):
                page.ele('x://i[@class="passIcon"]').click()

            while not page.ele('x://a[@class="uname e_icon at"]',timeout=search_timeout):
                try: page = page.latest_tab
                except: pass
                time.sleep(1)
            print("login end")

        page.get(self.query)


        while True:
            time.sleep(0.25)
            # 屏蔽滑动验证
            try: page.run_js('Object.defineProperties(navigator,{webdriver:{get:()=>false}})')
            except: pass

            box = page.ele('x://div[@class="joblist"]').children(timeout=search_timeout)
            for card in box:
                continue_flag = False
                # 职位
                job_name = card.ele('x://span[contains(@class,"jname")]',timeout=search_timeout).text
                for each in self.ignore_job:
                    if each in job_name:
                        continue_flag = True
                        break
                if continue_flag: continue

                # 薪资
                job_salary = card.ele('x://span[contains(@class,"sal")]',timeout=search_timeout).text
                job_salary = f'[{job_salary}]'
                for each in self.ignore_salary:
                    if each in job_salary:
                        continue_flag = True
                        break
                if continue_flag: continue

                # 学历
                job_tags = ''
                try:
                    for each in card.ele('x://div[@class="tags"]',timeout=search_timeout).children():
                        job_tags = f'{job_tags}/{each.attr("title")}'
                except: pass

                for each in self.ignore_edu:
                    if each in job_tags:
                        continue_flag = True
                        break
                if continue_flag: continue

                # 地区
                job_area = ''
                try: job_area += card.ele('x://div[@class="area"]',timeout=search_timeout).child(timeout=search_timeout).text
                except: pass
                for each in self.ignore_area:
                    if each in job_area:
                        continue_flag = True
                        break
                if continue_flag: continue

                job_company = card.ele('x://a[contains(@class,"cname")]',timeout=search_timeout).text
                job_company_tags = []
                try:
                    for idx,each in enumerate(card.ele('x://div[@class="bl"]',timeout=search_timeout).children(timeout=search_timeout)):
                        if 0 == idx or 1 == idx: continue
                        else: job_company_tags.append(each.text)
                except: pass

                try:
                    job_content = ''
                    job_link = ''
                    card.ele('x://div[@class="joblist-item-top"]',timeout=search_timeout).click()
                    tab = page.latest_tab
                    job_link = tab.url

                    try:
                        try:
                            # 滑动框验证
                            if not tab.ele('.bmsg job_msg inbox',timeout=1):
                                random_number = random.randint(50, 180)
                                ac = Actions(tab)
                                ac.hold('x://span[@class="nc_iconfont btn_slide"]')
                                ac.move(offset_x=random_number,duration=random.random() * 0.7 + 0.3)
                                time.sleep(random.random() * 0.7 + 0.3)
                                ac.move(offset_x=random.randint(300, 350) - random_number, duration=random.random() * 0.7 + 0.3)
                                ac.release()
                        except: pass

                        time.sleep(0.5)
                        tab.scroll.down(700)
                        time.sleep(1)
                        try:
                            for each in tab.eles('x://span[@class="sp4"]',timeout=search_timeout):
                                job_content = f'{job_content}/{each.raw_text}'
                        except: pass

                        if job_content: job_content += '\n'
                        try: job_content += tab.ele('x://div[@class="bmsg job_msg inbox"]',timeout=search_timeout).raw_text
                        except: pass

                        for each in self.ignore_job_require:
                            if each in job_content:
                                continue_flag = True
                                break

                        if continue_flag:
                            tab.close()
                            continue

                        print(f'职位:{job_name},工资:{job_salary},公司名:{job_company},位置:{job_area},详细内容:\n{job_content},招聘页面:{job_link}\n')
                        tab.close()

                        wb = load_workbook(self.file_path)
                        ws = wb.active
                        row = ws.max_row + 1

                        ws.cell(row=row, column=1, value=job_name)
                        ws.cell(row=row, column=2, value=job_salary)
                        ws.cell(row=row, column=3, value=job_company)
                        ws.cell(row=row, column=4, value=job_area)
                        ws.cell(row=row, column=5, value=job_content)
                        ws.cell(row=row, column=6, value=job_link)

                        wb.save(self.file_path)
                        wb.close()
                    except: pass
                    tab.close()
                except: pass

            next_page = page.ele('x://button[@class="btn-next"]')
            if 'disabled' == next_page.attr('disabled'):
                print('51job前程无忧 Crwal End')
                break
            next_page.click()
            time.sleep(0.5)

        page.close()


if __name__ == '__main__':
    try:
        _51job = _51Job(
            job_name='C++',
            job_salary='07',
            job_experience='01',
            job_degree='04',
            ignore_job=['教师', '老师', '销售', '保险', '消防', '前台', '客服', '营销', '售后', '管培生', '助理','经理', '硕士', '博士', '商务'],
            ignore_edu=['急聘', '1-3年', '3-5年', '5-10年', '2个月', '1年以下', '3年以上', '10年以上'],
            ignore_area=['北京', '上海'],
            ignore_salary=['面议', '[0-', '[1千-', '[2千-', '[3千-', '[4千-', '[5千-', '[6千-', '[7千-', '[8千-', '[9千-', '100-250元/天', '200-300元/天'],
            ignore_job_require=['硕士学历', '硕士及以上学历', '硕士以上学历', '博士学历', '博士及以上学历', '博士以上学历', '985', '211', '出差', '重点大学'],
            login_in=True,
            overlap=False
        )
        _51job.run()
    except: os.kill(os.getpid(), signal.SIGINT)