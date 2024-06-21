import sys

from DrissionPage import ChromiumPage
from openpyxl import load_workbook, Workbook
from urllib.parse import quote

import time
import os
import signal

# 薪资 - salary
"""
不限：''
3k以下：'0-3000'
3-5k：'3000-5000'
5-10k：'5000-10000'
10-20k：'10000-20000'
20-30k：'20000-30000'
30-40k：'30000-40000'
40-50k：'40000-50000'
50-60k：'50000-60000'
60k以上：'60000-1000000000'
"""

# 学历 - job_degree
"""
不限：''
本科：'3'
硕士：'4'
博士：'5'
其他：'7'
"""

class BoLe():
    def __init__(self,job_name='',job_salary='',job_degree='',
        ignore_job=[],ignore_edu=[],ignore_area=[],ignore_salary=[],ignore_job_require=[],
        overlap=False,login_in=False,username='',password=''):
        self.username = username
        self.password = password
        self.login_in = login_in
        self.ignore_job = ignore_job
        self.ignore_edu = ignore_edu
        self.ignore_area = ignore_area
        self.ignore_salary = ignore_salary
        self.ignore_job_require = ignore_job_require

        self.query = f'https://www.bolexiaozhao.com/index.php/Home/Hotrec/index/type/1'
        if job_degree: self.query = f'{self.query}/edu_id/{job_degree}'
        if job_salary: self.query = f'{self.query}/price_id/{job_salary}'
        if job_name: self.query = f'{self.query}/key/{quote(job_name)}'
        self.query = f'{self.query}.html'
        file_name = '../output/bole_crawl_output.xlsx'

        if not os.path.exists('../output'): os.mkdir('../output')
        if overlap:
            counter = 1
            while os.path.exists(file_name):
                file_name = f'../output/bole_crawl_output_{counter}.xlsx'
                counter += 1
        elif os.path.exists(file_name):
            os.remove(file_name)

        self.file_path = file_name

        wb = Workbook()
        ws = wb.active
        row = ws.max_row

        ws.cell(row=row, column=1, value='职位')
        ws.cell(row=row, column=2, value='工资')
        ws.cell(row=row, column=3, value='公司名')
        ws.cell(row=row, column=4, value='位置')
        ws.cell(row=row, column=5, value='详细内容')
        ws.cell(row=row, column=6, value='截至时间')
        ws.cell(row=row, column=7, value='招聘页面')

        wb.save(self.file_path)
        wb.close()

    def pre_num(self,str):
        num = ['0','1','2','3','4','5','6','7','8','9']
        result = ''

        for each in str:
            if each in num: result += each
            else: break

        return  result

    def run(self):
        # search_timeout必需大于0.2
        search_timeout = 0.25
        
        page = ChromiumPage()

        # 账号密码登录
        if self.login_in:
            print("login.....")
            page.get('https://www.bolexiaozhao.com/')

            if not page.ele('x://b[@class="user-name"]',timeout=search_timeout):
                if '' == self.username or '' == self.password:
                    print('username is empty or password is empty')
                    sys.exit()
                page.ele('x://a[@class="btn-login"]', timeout=search_timeout).click()
                page.ele('x://input[@name="login-phone"]', timeout=search_timeout).input(vals=self.username)
                page.ele('x://input[@name="login-password"]', timeout=search_timeout).input(vals=self.password)
                page.ele('x://input[@name="login-remember"]', timeout=search_timeout).click()
                page.ele('x://a[contains(@class,"btn-login-submit")]', timeout=search_timeout).click()
                time.sleep(2)
            print("login end")

        page.get(self.query)

        while True:
            time.sleep(0.25)
            # 多此一举，屏蔽滑动验证
            try: page.run_js('Object.defineProperties(navigator,{webdriver:{get:()=>false}})')
            except: pass

            continue_flag = False

            box = page.ele('x://ul[@class="job-list"]').children(timeout=search_timeout)
            for card in box:
                # 职位
                job_list = ''
                for each in card.ele('x://div[@class="company-position"]',timeout=search_timeout).children(timeout=search_timeout):
                    job_list = f'{job_list}/{each.text}'

                for each in self.ignore_job:
                    if each in job_list:
                        continue_flag = True
                        break
                if continue_flag: continue

                job_num = self.pre_num(card.ele('x://span[@class="position-num"]',timeout=search_timeout).text)
                if '' == job_num or '0' == job_num: continue

                job_company = card.ele('x://h3[@class="text-line1"]',timeout=search_timeout).text
                job_area = card.ele('x://span[@class="addr"]',timeout=search_timeout).text

                try:
                    if not job_list: continue

                    tab = card.ele('x://a[@class="float-r btn btn-ljws"]',timeout=search_timeout).click.middle()
                    page.set.tab_to_front(tab)

                    try:
                        for each in tab.ele('x://ul[contains(@class,"position-list")]').children(timeout=search_timeout):
                            job_name_label = each.ele('x://span[contains(@class,"position-name")]',timeout=search_timeout).child(timeout=search_timeout)

                            job_name = job_name_label.text
                            if job_name not in job_list: continue

                            job_time = each.ele('x://p[@class="time"]',timeout=search_timeout).text
                            job_content = ''

                            # 薪资
                            job_salary = each.ele('x://span[contains(@class,"position-reward")]',timeout=search_timeout).text
                            job_salary = f'[{job_salary}]'
                            for each in self.ignore_salary:
                                if each in job_salary:
                                    continue_flag = False
                                    break

                            if not continue_flag:
                                time.sleep(0.25)
                                new_tab = job_name_label.click.middle()
                                page.set.tab_to_front(new_tab)

                                time.sleep(0.5)
                                new_tab.scroll.down(400)
                                time.sleep(1)

                                job_link = job_name_label.attr('href')
                                information = new_tab.ele('x://div[contains(@class,"left-box")]').children(timeout=search_timeout)
                                for p in information: job_content = f'{job_content}{p.raw_text}\n'
                                for each in self.ignore_job_require:
                                    if each in job_content:
                                        continue_flag = True
                                        break

                                if continue_flag:
                                    new_tab.close()
                                    continue

                                new_tab.close()
                                print(f'职位:{job_name},工资:{job_salary},公司名:{job_company},位置:{job_area},详细内容:\n{job_content},截至时间：{job_time},招聘页面:{job_link}\n')

                                wb = load_workbook(self.file_path)
                                ws = wb.active
                                row = ws.max_row + 1

                                ws.cell(row=row, column=1, value=job_name)
                                ws.cell(row=row, column=2, value=job_salary)
                                ws.cell(row=row, column=3, value=job_company)
                                ws.cell(row=row, column=4, value=job_area)
                                ws.cell(row=row, column=5, value=job_content)
                                ws.cell(row=row, column=6, value=job_time)
                                ws.cell(row=row, column=7, value=job_link)

                                wb.save(self.file_path)
                                wb.close()
                    except: pass
                    tab.close()
                except: pass

            next_page = page.ele('x://a[@class="next"]')
            if 'javascript:;' == next_page.attr('href'):
                print('伯乐校招 Crwal End')
                break
            next_page.click()
            time.sleep(0.5)
        page.close()


if __name__ == '__main__':
    try:
        bole = BoLe(
            job_name='',
            job_salary='5000-10000',
            job_degree='3',
            ignore_job=['教师', '老师', '销售', '保险', '消防', '前台', '客服', '营销', '售后', '管培生', '助理','经理', '硕士', '博士', '商务'],
            ignore_edu=['急聘', '1-3年', '3-5年', '5-10年', '2个月', '1年以下', '3年以上', '10年以上'],
            ignore_area=['北京', '上海'],
            ignore_salary=['面议', '[0-', '[1000-', '[2000-', '[3000-', '[4000-', '[5000-', '[6000-', '[7000-', '[8000-', '[9000-', '100-250元/天', '200-300元/天'],
            ignore_job_require=['硕士学历', '硕士及以上学历', '硕士以上学历', '博士学历', '博士及以上学历', '博士以上学历', '985', '211', '出差'],
            login_in=True,
            username='', # 手机号
            password='',
            overlap=False
        )
        bole.run()
    except: os.kill(os.getpid(), signal.SIGINT)